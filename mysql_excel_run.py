import openpyxl
import pymysql


class DatabaseClearProcess:
    """数据库执行脚本
        path：xlsx文件地址
        model：运行模式（DEBUG/ONLINE）
    """

    def __init__(self, path, model):
        self.path = path
        self.model = model

    # 通过xlsx地址获取workbook对象
    def do(self):
        wb = openpyxl.load_workbook(self.path)
        self.sheet_of_database(wb)
        wb.save(self.path)

    # 循环遍历所有sheet，sheet名为数据库名
    def sheet_of_database(self, wb):
        for sheet_of_database in list(wb.sheetnames):
            self.process_sheet(wb[sheet_of_database], sheet_of_database)

    # 检查sheet是否合法
    def sheet_check(self, sheet):
        rows = list(sheet.rows)

        # 当行数小于等于4行，则属于无效Sheet页
        if len(rows) < 4:
            return False

        # 当前sheet表名和执行SQL所在下标
        table_index = self.table_index(sheet)
        sql_index = self.sql_index(sheet)

        if table_index is None or sql_index is None:
            return False

        return True

    # 循环遍历所有表，并执行响应SQL语句
    def process_sheet(self, sheet, database_name):
        if self.sheet_check(sheet) is False:
            print(f"Sheet：{database_name}，不符合约定规范，请检查！\n")
            return

        # 当前sheet表名和执行SQL所在下标
        table_index = self.table_index(sheet)
        sql_index = self.sql_index(sheet)
        status_index = self.status_index(sheet)

        try:
            url, port, username, password = self.read_jdbc_resource(sheet)
        except SystemExit:
            print(f"Sheet：{database_name}，数据库配置缺失，请检查！\n")
            return

        db = self.open_db(url, port, username, password, database_name)

        if db is None:
            return False

        # 从第三行开始读取数据行
        all_rows = list(sheet.rows)[3:]

        for index, row in enumerate(all_rows):
            table_name = row[table_index].value
            sql = row[sql_index].value

            if table_name is None or table_name == "":
                continue

            if sql is None or sql == "":
                continue

            print("[SQL 开始]执行命令参数：\n"
                  "url：%s，port：%s，username：%s，password：%s，db：%s，table：%s\n"
                  "sql: %s\n-------------------------" % (url, port, username, password, database_name, table_name, sql))

            status = self.run(db, sql)

            print(f"[SQL 结束]执行完成，执行状态：{status} \n")
            # excel下标从1开始，Python数组下标从0开始，所以需要整体+1
            # 又因为前三行属于基础配置，不是表，所以需要跳过+3行
            sheet.cell(row=index + 1 + 3, column=status_index + 1, value=str(status))

        db.close()

    def table_index(self, sheet):
        return self.list_of_content(list(sheet.rows)[2], "表名")

    def sql_index(self, sheet):
        return self.list_of_content(list(sheet.rows)[2], "执行SQL")

    def status_index(self, sheet):
        return self.list_of_content(list(sheet.rows)[2], "完成状态")

    def url_index(self, sheet):
        return self.list_of_content(list(sheet.rows)[0], "数据库连接地址")

    def port_index(self, sheet):
        return self.list_of_content(list(sheet.rows)[0], "数据库端口号")

    def username_index(self, sheet):
        return self.list_of_content(list(sheet.rows)[0], "用户名")

    def password_index(self, sheet):
        return self.list_of_content(list(sheet.rows)[0], "密码")

    @staticmethod
    def list_of_content(list_data, str_content):
        for index, content in enumerate(list_data):
            if content.value == str_content:
                return index

    def read_jdbc_resource(self, sheet):
        url = self.url_index(sheet)
        port = self.port_index(sheet)
        username = self.username_index(sheet)
        password = self.password_index(sheet)

        if url is None or port is None or username is None or password is None:
            raise SystemExit("数据库参数配置异常！")

        url = sheet.cell(row=2, column=url + 1)
        port = sheet.cell(row=2, column=port + 1)
        username = sheet.cell(row=2, column=username + 1)
        password = sheet.cell(row=2, column=password + 1)
        return url.value, port.value, username.value, password.value

    @staticmethod
    def open_db(url, port, username, password, database):
        try:
            db = pymysql.connect(
                host=url,
                port=int(port),
                user=username,
                passwd=str(password),
                db=database,
                charset='utf8'
            )
            return db
        except Exception as e:
            print("数据库连接失败：Fail。\n失败原因：%s \n"
                  "url: %s，port：%s，username：%s，password：%s，db：%s\n" % (e, url, port, username, password, database))
            return None

    @staticmethod
    def exec(db, sql):
        cur = db.cursor()
        # 多条语句通过"|"分割。
        all_sql = sql.split("|")

        try:
            for sql in all_sql:
                affected = cur.execute(sql)
                print(f"[{sql}]，受影响的行数数目：" + str(affected))
            db.commit()
            return True
        except Exception as e:
            db.rollback()
            print("[SQL 任务] SQL执行失败，原因：%s" % e)
            return False
        finally:
            cur.close()

    # 运行表SQL
    def run(self, db, sql):
        if self.model == "ONLINE":
            print("执行状态：Running...")
            status = self.exec(db, sql)
            print("执行状态：Completed")

            return status
        elif self.model == "DEBUG":
            return True
        else:
            print("当前运行模式异常：" + self.model)
            return False


if __name__ == '__main__':
    file_path = input("请输入excel/xlsx文件地址:" + '\n')
    operator = DatabaseClearProcess(file_path, "ONLINE")
    operator.do()
